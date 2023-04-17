# Imports ---------------------------------------------------------------------
import os, sys, time, logging, platform, importlib, datetime
import openpyxl, inquirer, yaml, pyvisa
import numpy as np
import pyqtgraph as pg
from openpyxl.drawing.image import Image
from art import text2art
from rich.console import Console
from rich.traceback import install
from rich.logging import RichHandler
from rich.progress import Progress
from rich.panel import Panel
from PyQt6.QtCore import QTimer
from PyQt6.QtWidgets import QApplication,QMainWindow, QStyleFactory, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QTabWidget, QGridLayout



# Global Variables ------------------------------------------------------------

ProgramInfo = {
    'name': 'PyCharMem',
    'version': '0.0.2',
    'author': 'Ricardo E. Silva',
    'email': 'ricardoedgarsilva@tecnico.ulisboa.pt',
    'description': 'PyCharMem is a Python program that allows you to measure the charge memory of a device.',
    'license': 'GNU General Public License v3.0',
    'url': 'https://github.com/ricardoesilva/PyCharMem'  
}

#Common functions -------------------------------------------------------------
def verbose_debug(verbose:bool):
    """Verbose Debug"""
    # Verbose Debug
    if verbose:
        logging.basicConfig(level="NOTSET", format="%(message)s", datefmt="[%X]", handlers=[RichHandler()])
    else:
        logging.basicConfig(level='INFO', format="%(message)s", datefmt="[%X]", handlers=[RichHandler()])

    return logging.getLogger("rich")

def clear_terminal():
    """Clear Terminal"""
    os.system('cls||clear')

def printnewlines(lines:int):
    """Print New Lines"""
    print(lines*"\n")

def splash_screen(console:Console):
    """Splash Screen"""
    print(text2art(ProgramInfo['name']))
    console.print(f"Version: {ProgramInfo['version']}",style='bold blue')
    console.print(f"Author: {ProgramInfo['author']}",style='bold blue')
    printnewlines(5)

def config_open(logger:logging.Logger):
    """Open the config file"""
    try:
        match platform.system():
            case 'Windows': os.system(f'start config.yml')
            case 'Darwin':  os.system(f"open config.yml")
            case 'Linux':   os.system(f"open config.yml")
    except:
        logger.critical('Config file not found')
        quit()

def config_read(logger:logging.Logger):
    """Read the config file"""
    try:
        with open('config.yml', 'r') as file: config = yaml.safe_load(file)
        return config
    except:
        logger.critical('Config file not found')
        quit()

def path_exists(path:str):
    """Check if path exists"""
    return os.path.exists(path)

def get_path():
    """Get the path to the file"""
    #path = os.path.dirname(os.path.realpath(__file__)) This retuns the parent of /PyCharMem
    #return path[:path.rfind('\\')]

    return os.path.dirname(os.path.realpath(__file__))

def mkdir(logger:logging.Logger, path:str):
    '''Creates the path if it doesn't exist'''
    if path_exists(path): 
        logger.debug(f'Folder {path} already exists')
    else: 
        os.makedirs(path)
        logger.debug(f'Folder {path} created')

def get_index(path:str):
    '''Returns the index of the file in the folder'''
    return len(os.listdir(path))

def get_filename(path:str ,sample:str, device:str):
    '''Returns the filename with the date_sample_device_index.xlsx format'''
    index = get_index(path)
    date = datetime.datetime.now().strftime('%Y%m%d')
    return f'{date}_{sample}_{device}_{index}.xlsx'

def get_available_addresses():
    '''Returns the available adresses'''
    rm = pyvisa.ResourceManager()
    addresses = rm.list_resources()
    return addresses

def print_available_addresses(logger:logging.Logger, console:Console, addresses:list):
    '''Prints the available adresses'''

    printnewlines(1)
    console.print(Panel.fit("[bold]Available Addresses[/bold]", border_style="green"))    
    if len(addresses) == 0: logger.critical('No addresses found!')
    else:
        for address in addresses:
            console.print(address)

def check_address(logger:logging.Logger, addresses:list, address:str):
    '''Checks if the address is in the list of available addresses'''
    if address not in addresses: 
        logger.critical('Sourcemeter address in config file not found! Please edit config file with correct address')
        quit()

def check_missing_params(logger:logging.Logger, my_dict:dict, my_list:list):
    '''Checks if the dictionary has all the keys in the list'''
    
    missing_items = [item for item in my_list if item not in my_dict]

    if len(missing_items) == 0: logger.info('All measurement parameters found!')
    else:
        logger.critical('Missing measurement parameters:')
        for item in missing_items:
            logger.critical(f'- {item}')
        quit()

def create_list(logger:logging.Logger, condition_values:list):
    '''Creates a list of values'''
    [cycle,max,min,step] = condition_values
    # Create list of values from 0 to max
    listp_oneway = np.arange(0,max,step)
    listn_oneway = np.arange(0,min,-step)
    # Create list of values from max to 0 and from 0 to min
    listp = np.concatenate((listp_oneway,np.array([max]),np.flip(listp_oneway)))
    listn = np.concatenate((listn_oneway,np.array([min]),np.flip(listn_oneway)))

    try:
        match cycle:
            case '+': return listp
            case '-': return listn
            case '+-': return np.concatenate((listp,listn))
            case '-+': return np.concatenate((listn,listp))
    except:
        logger.critical('Invalid cycle type!')
        quit() 

def get_datetime():
    '''Returns date and time in the format YYYY-MM-DD HH:MM:SS:MS'''
    return datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S:%f')

def import_module(logger:logging.Logger, type:str, inst_type=None,measurement_type=None,plot_type=None):
    '''Imports a module file from string'''

    logger.debug(f'Importing {type} module')

    match type:
        case 'instrument':
            file_name = f'instruments.{inst_type}'
            obj_name = 'Instrument'
        
        case 'measurement':
            file_name = f'measurements.{measurement_type}'
            obj_name = 'Measurement'

        
        case 'plot':
            file_name = f'plots.{plot_type}'
            obj_name = 'Plots'

        
    file = importlib.import_module(file_name)
    logger.debug(f'{file_name} module imported')
    obj = getattr(file, obj_name)
    logger.debug(f'{obj_name} class imported')
    return obj
    
def menu(logger:logging.Logger, console:Console, type:str):
    '''Menu function'''
    
    logger.debug(f'Menu called! Type: {type}')
    match type:
        case 'main':
            name = 'Main Menu'
            message = 'Select an option:'
            choices = ['Exit','Select Measurement','Print Available Addresses','Edit Configuration']
        
        case 'measurements':
            try:
                list_dir = os.listdir('measurements')
                measurement_types = [filename.split('.')[0] for filename in list_dir if filename.endswith('.py')]
                logger.debug(f'{len(measurement_types)} measurement types found')
            except:
                logger.critical('No measurement types found!')
                time.sleep(2)
                quit()

            name = 'Measurement Selection Menu'
            message = 'Select a measurement:'
            choices = measurement_types
            choices.append('Back')

    printnewlines(1)
    console.print(Panel.fit(f"[bold]{name}[/bold]", border_style="green"))
    menu = [inquirer.List('choice',message=message,choices=choices)]
    answer = inquirer.prompt(menu)
    logger.debug(f'Returning answer: {answer}')
    return answer['choice']

def ask_for_comment(logger:logging.Logger):
    comment = [inquirer.Text('comment',message="What would you like to comment?",validate=lambda _, x: len(x) > 0),]
    answer = inquirer.prompt(comment)
    return answer['comment']


# Important Classes --------------------------------------------------------------
class FileSave:
    def __init__ (self, logger:logging.Logger, sample:str, device:str):
        '''Initializes the class'''

        #Assign path and make folder if it doesn't exist
        self.path = os.path.join(get_path(), "data", sample, device)
        mkdir(logger,self.path)
        #Create file
        self.file_name = get_filename(self.path,sample,device)
        self.file_path = os.path.join(self.path, self.file_name)
        self.wb = openpyxl.Workbook()
        self.wb.create_sheet('config')
        self.wb.create_sheet('data')
        self.wb.create_sheet('plots')
        self.wb.remove(self.wb['Sheet'])
        self.wb.save(self.file_path)

    def save_config(self, logger:logging.Logger, config:dict, measurement_type:str):
        '''Saves the config file in the config sheet'''
        
        #Open config sheet and set title
        self.ws = self.wb['config']
        
        #read config file
        sections = ['sample','instrument', measurement_type]
        for section in sections:
            self.ws.append([section])
            for key, value in config.get(section).items():
                self.ws.append([key,value])
            self.ws.append([])
        self.wb.save(self.file_path)
        logger.debug('Configuration file saved in config sheet')

    def save_headers(self, logger:logging.Logger, headers_list:list):
        self.ws = self.wb['data']
        for i in range(len(headers_list)):
            self.ws.cell(row=1, column=i+1).value = headers_list[i]
        self.wb.save(self.file_path)
        logger.debug('Headers saved in data sheet')

    def save_result(self, logger:logging.Logger, result:list):
        self.ws = self.wb['data']
        row = self.ws.max_row + 1
        for i in range(len(result)):
            self.ws.cell(row=row, column=i+1).value = result[i]
        self.wb.save(self.file_path)
        logger.debug('result saved in data sheet')
    
    def save_plots(self, logger:logging.Logger):
        self.ws = self.wb['plots']
        img = Image('temp/plot.png')
        self.ws.add_image(img, 'B2')
        self.wb.save(self.file_path)
        os.remove('temp/plot.png')
        logger.debug('Plot saved in plots sheet')

class Logbook:
    def __init__(self, logger:logging.Logger, sample:str):
        '''Initializes the class'''
        #Assign path and make folder if it doesn't exist
        self.path = os.path.join(get_path(), "data", sample)
        mkdir(logger,self.path)
        #Create file
        self.file_path = os.path.join(self.path, 'logbook.xlsx')
        
        #Check if file exists
        if path_exists(self.file_path):
            self.wb = openpyxl.load_workbook(self.file_path)
            logger.debug('Logbook file already exists')

        else:
            self.wb = openpyxl.Workbook()
            self.wb.create_sheet('logbook')
            self.wb.remove(self.wb['Sheet'])

            #Create headers
            self.ws = self.wb['logbook']
            self.ws.title = 'logbook'
            headers_list = ['Date','File','Comment']
            for i in range(len(headers_list)):
                self.ws.cell(row=1, column=i+1).value = headers_list[i]
            
            self.wb.save(self.file_path)
            logger.debug('Logbook file created')

    def save_log(self, logger: logging.Logger, date:str, file:str, comment:str):
        '''Saves the log in the logbook sheet'''
        self.ws = self.wb['logbook']
        row = self.ws.max_row + 1
        self.ws.cell(row=row, column=1).value = date
        self.ws.cell(row=row, column=2).value = file
        self.ws.cell(row=row, column=3).value = comment
        self.wb.save(self.file_path)
        logger.debug('Log saved in logbook sheet')

class MeasurementWindow(QMainWindow):
    def __init__(self,meas):
        super().__init__()
        self.setWindowTitle(meas.name)
        self.setGeometry(100,100,800,600)

        #Create a tab widget
        self.tab_widget = QTabWidget(self)

        #Create a Plot tab and add it to the tab widget
        plot_tab = QWidget()
        plot_layout = QVBoxLayout(plot_tab)
        plot_grid = QGridLayout()
        [rows,cols] = meas.plot_grid
        for row in range(rows):
            for col in range(cols):
                plot_widget = pg.PlotWidget()
                plot_widget.setTitle(meas.plot_titles[row][col])
                plot_widget.setLabel('bottom',meas.plot_labels[row][col][0])
                plot_widget.setLabel('left',meas.plot_labels[row][col][1])
                plot_grid.addWidget(plot_widget,row,col)

        plot_layout.addLayout(plot_grid)
        plot_tab.setLayout(plot_layout)
        self.tab_widget.addTab(plot_tab, "Plots")


        #Create a tab and add it to the tab widget
        data_tab = QWidget()
        data_layout = QVBoxLayout(data_tab)
        self.data_table = QTableWidget()
        self.data_table.setRowCount(0)
        self.data_table.setColumnCount(len(meas.headers))
        self.data_table.setHorizontalHeaderLabels(meas.headers)
        data_layout.addWidget(self.data_table)
        data_tab.setLayout(data_layout)
        self.tab_widget.addTab(data_tab, "Data")

        #Ad the widget to the main window
        self.setCentralWidget(self.tab_widget)

        #Call measure function

    




# Main Function ------------------------------------------------------------------

def main():
    install()
    console = Console()
    logger = verbose_debug(False)
    running = True

    #Splash Screen
    clear_terminal()
    splash_screen(console)

    #Load config file
    config = config_read(logger)
    logger.info('Configuration file loaded')

    #Print Available Addresses
    adresses = get_available_addresses()
    print_available_addresses(logger,console,adresses)

    #Load Instruments
    inst_class = import_module(logger=logger,type='instrument',inst_type=config.get('instrument').get('model'))
    inst = inst_class(logger,config)
    logger.info('Instrument loaded')

    while running:
        option = menu(logger,console,'main')

        match option:
            case 'Exit':
                running = False
                logger.info('Exiting program')
            
            case 'Print Available Addresses':
                print_available_addresses(logger,console,get_available_addresses())

            case 'Edit Configuration':
                config_open(logger)
            
            case 'Select Measurement':
                #Load Measurements menu and go back to main menu if 'Back' is selected
                option2 = menu(logger,console,'measurements')
                if option2 == 'Back': continue

                #Import measurement class
                meas_class = import_module(logger=logger,type='measurement',measurement_type=option2)
                meas = meas_class(logger,config,inst)

                #Check if all measurements are present
                check_missing_params(logger,meas.params[meas.name],meas.nparams)
                logger.info('All parameters present')

                filesave = FileSave(logger,config.get('sample').get('name'),config.get('instrument').get('model'))
                filesave.save_config(logger,config,option2)
                filesave.save_headers(logger,meas.headers)
                logbook = Logbook(logger,config.get('sample').get('name'))

                #Import plot class
                n_cycles = int(config.get(meas.name).get('n_cycles'))

                app = QApplication(sys.argv)
                app.setStyle(QStyleFactory.create("Fusion"))
                app.setStyleSheet("QMainWindow::title {background-color: #333333;}")
                window = MeasurementWindow(meas)
                window.show()
                app.exec()
                continue









if __name__ == "__main__":
    main()
