import os
import openpyxl
from modules.common import get_filename, get_path, mkdir

class FileSaver:
    def __init__(self,logger, sample, device):
        self.path = os.path.join(get_path(), "data", sample, device)
        self.logbook_path = os.path.join(get_path(), "data", sample,'logbook.xlsx')
        mkdir(logger,self.path)
        self.file_name = get_filename(self.path,sample,device)
        self.file_path = os.path.join(self.path, self.file_name)

        #debug
        logger.debug(f'Data file path: {self.file_path}')
        logger.debug(f'Logbook file path: {self.logbook_path}')

        #create excel file
        self.wb = openpyxl.Workbook()
        self.wb.create_sheet('config')
        self.wb.create_sheet('data')
        self.wb.create_sheet('plots')
        self.wb.remove(self.wb['Sheet'])
        self.wb.save(self.file_path)

        #create logbook
        self.logbook = openpyxl.Workbook()
        try:
            self.logbook = openpyxl.load_workbook(self.logbook_path)
            logger.debug('Logbook loaded')
        except:
            self.logbook.create_sheet('logbook')
            self.logbook.remove(self.logbook['Sheet'])
            self.logbook.save(self.logbook_path)
            logger.debug('Logbook created')
        
    def save_config(self,logger,config,measurement_type):
        #create excel sheet
        self.ws = self.wb['config']
        self.ws.title = 'config'

        #read config file
        sections = ['sample','sourcemeter',f'measurement/{measurement_type}']
        for section in sections:
            self.ws.append([section])
            for key, value in config.items(section):
                self.ws.append([key,value])
            self.ws.append([])
        self.wb.save(self.file_path)
        logger.debug('Config sheet saved')

    def save_headers(self, headers_list):
        self.ws = self.wb['data']
        self.ws.title = 'data'
        for i in range(len(headers_list)):
            self.ws.cell(row=1, column=i+1).value = headers_list[i]
        self.wb.save(self.file_path)

    def save_result(self,result):
        self.ws = self.wb['data']
        row = self.ws.max_row + 1
        for i in range(len(result)):
            self.ws.cell(row=row, column=i+1).value = result[i]
        self.wb.save(self.file_path)
    
    def save_plots(self,image):
        self.ws = self.wb['plots']
        self.ws.add_image(image, 'B1')
        self.wb.save(self.file_path)
