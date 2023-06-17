# Imports
import os
import sys
import time
import atexit
import logging
import platform
import importlib
import datetime
import openpyxl
import inquirer
import yaml
import pyvisa
import numpy as np
import pyqtgraph as pg
from art import text2art
from rich.console import Console
from rich.traceback import install
from rich.logging import RichHandler
from rich.progress import Progress
from rich.panel import Panel
from PyQt6.QtCore import QThread, pyqtSignal
from PyQt6.QtWidgets import (QApplication, QMainWindow, QStyleFactory, QTableWidget,
                              QTableWidgetItem, QVBoxLayout, QWidget, QTabWidget, QGridLayout)

# Global Variables
PROGRAM_INFO = {
    'name': 'PyCharMem',
    'version': '0.0.5',
    'author': 'Ricardo E. Silva',
    'email': 'ricardoedgarsilva@tecnico.ulisboa.pt',
    'description': 'PyCharMem is a Python program that allows you to measure the charge memory of a device.',
    'license': 'GNU General Public License v3.0',
    'url': 'https://github.com/ricardoedgarsilva/PyCharMem'
}

def verbose_debug(verbose):
    """
    Configure a verbose debugging setup.

    Initializes a logging configuration with customizable verbosity based on the provided `verbose` parameter.
    If `verbose` is True, the log level is set to "NOTSET"; otherwise, it is set to 'INFO'.
    The log format displays only the log message with a timestamp in the format of [hour:minute:second].
    The log output is handled by a RichHandler.

    Returns:
        logging.Logger or None: The logger object from the "rich" logger if the setup is successful.
        If an exception occurs during setup, prints the error message and terminates the program, returning None.
    """

    try:
        log_level = "NOTSET" if verbose else 'INFO'
        logging.basicConfig(level=log_level, format="%(message)s", datefmt="[%X]", handlers=[RichHandler()])
        return logging.getLogger("rich")
    except Exception as e:
        print(f"Error: {e}")
        quit()
        return None

def clear_terminal():
    """
    Clear the terminal screen.

    Uses the appropriate system command based on the operating system.
    If running on Windows, executes the 'cls' command.
    If running on Unix-based systems, executes the 'clear' command.
    """
    
    try:
        os.system('cls||clear')
    except Exception:
        print("Unable to clear terminal screen. Please clear the screen manually.")

def print_newlines(lines):
    """
    Print a specified number of newline characters.

    Args:
        lines (int): The number of newline characters to be printed.
    """

    try:
        print(lines * "\n")
    except TypeError:
        print("Invalid argument.")

def splash_screen(console):
    """
    Display a splash screen with program information.

    Prints the program name as ASCII art using the 'text2art' function.
    Displays the program version, repository URL, and author using the 'console.print' function.
    Inserts 5 newlines using the 'print_newlines' function.

    Args:
        console (Console): A console object for rich text output.
    """
        
    print(text2art(PROGRAM_INFO['name']))
    console.print(f"Version: {PROGRAM_INFO['version']}", style='bold blue')
    console.print(f"Repositoty: {PROGRAM_INFO['url']}", style='bold blue')
    console.print(f"Author: {PROGRAM_INFO['author']}", style='bold blue')
    print_newlines(5)

def open_config(logger):
    """
    Open the config file using the appropriate system command.

    Attempts to open the 'config.yml' file based on the operating system.
    On Windows, uses the 'start' command.
    On macOS, uses the 'open' command.
    On Linux, uses the 'open' command.
    If the config file is not found, logs a critical message using the provided logger and exits the program.

    Args:
        logger (logging.Logger): The logger object to use for logging error messages.
    """


    try:
        os.system({'Windows': 'start config.yml', 'Darwin': 'open config.yml', 'Linux': 'open config.yml'}.get(platform.system(), ''))
    except:
        logger.critical('Config file not found')
        quit()

def read_config(logger):
    """
    Read and load the configuration from a YAML file.

    Attempts to read the 'config.yml' file and load its contents using the `yaml.safe_load` function.
    If the file is successfully loaded, logs an informational message using the provided logger and returns the configuration.
    If the file is not found, logs a critical message and exits the program.
    If the file is not a valid YAML file, logs a critical message and exits the program.

    Args:
        logger (logging.Logger): The logger object to use for logging error messages.

    Returns:
        dict: The configuration loaded from the YAML file.
    """
    
    try:
        with open('config.yml', 'r') as file:
            config = yaml.safe_load(file)
        logger.info('Config file loaded!')
        return config
    except FileNotFoundError:
        logger.critical('Config file not found')
        quit()
    except yaml.YAMLError:
        logger.critical('Config file is not a valid YAML file')
        quit()

def path_exists(path):
    """
    Check if a path exists.

    Checks whether the specified `path` exists in the file system.
    Returns `True` if the path exists, `False` otherwise.
    If an OSError occurs during the check, it returns `False`.

    Args:
        path (str): The path to be checked.

    Returns:
        bool: `True` if the path exists, `False` otherwise.
    """

    try:
        return os.path.exists(path)
    except OSError:
        return False

def get_path():
    """
    Get the path of the current file or module.

    If the script is being run as the main module (__name__ == '__main__'), returns the directory name
    of the current file after resolving symbolic links using `os.path.realpath`.
    Otherwise, returns the directory name of the current file using `os.path.abspath`.

    Returns:
        str: The path of the current file or module.
    """
    
    if __name__ == '__main__':
        return os.path.dirname(os.path.realpath(__file__))
    else:
        return os.path.dirname(os.path.abspath(__file__))

def mkdir(logger, path):
    """
    Create a directory at the specified path.

    Checks if the directory already exists using the `path_exists` function.
    If the directory does not exist, creates it using `os.makedirs`.
    Logs a debug message with the logger indicating whether the folder was created or already exists.
    If an OSError occurs during the creation, logs an error message with the logger.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        path (str): The path of the directory to be created.
    """

    try:
        if not path_exists(path):
            os.makedirs(path)
            logger.debug(f'Folder {path} created')
        else:
            logger.debug(f'Folder {path} already exists')
    except OSError as e:
        logger.error(f'Failed to create folder {path}: {e}')

def get_index(path):
    """
    Get the index of the last element in a directory.

    Retrieves the index by counting the number of items in the specified `path` using `os.listdir`.
    If the directory is not found, returns 0 as the index.

    Args:
        path (str): The path of the directory.

    Returns:
        int: The index of the last element in the directory.
    """   
    index = 0
    try:
        index = len(os.listdir(path))
    except FileNotFoundError:
        pass
    return index

def get_filename(path, sample, device):
    """
    Generate a filename for a new Excel file.

    Retrieves the index of the last element in the specified `path` using the `get_index` function.
    Gets the current date in the format '%Y%m%d' using `datetime.datetime.now().strftime`.
    Constructs and returns a filename using the format '{date}_{sample}_{device}_{index}.xlsx'.

    Args:
        path (str): The path of the directory where the file will be saved.
        sample (str): The name or identifier of the sample.
        device (str): The name or identifier of the device.

    Returns:
        str: The generated filename for the Excel file.
    """

    index = get_index(path)
    date = datetime.datetime.now().strftime('%Y%m%d')
    return f'{date}_{sample}_{device}_{index}.xlsx'

def get_available_addresses():
    """
    Get the available addresses of USB devices.

    Attempts to create a PyVISA resource manager using `pyvisa.ResourceManager`.
    Retrieves the list of available resource addresses using `rm.list_resources()`.
    If no USB devices are found, prints a message and returns an empty list.

    Returns:
        List[str]: A list of available addresses of USB devices.
    """

    try:
        rm = pyvisa.ResourceManager()
        addresses = rm.list_resources()
        return addresses
    except pyvisa.VisaIOError:
        print("No USB devices found")
        return []

def print_available_addresses(logger, console, addresses):
    """
    Print the available addresses of USB devices.

    Prints a blank line using the `print_newlines` function.
    If no addresses are provided, logs a critical message using the provided logger and returns.
    Prints a panel with the title "Available Addresses" using the `console.print` function.
    Prints each address from the provided list using the `console.print` function.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        console (Console): A console object for rich text output.
        addresses (List[str]): A list of available addresses of USB devices.
    """

    
    print_newlines(1)
    
    if len(addresses) == 0:
        logger.critical('No addresses found!')
        return
    
    console.print(Panel.fit("[bold]Available Addresses[/bold]", border_style="green"))
    
    for address in addresses:
        console.print(address)

def check_address(logger, addresses, address):
    """
    Check if the specified address is valid in the configuration.

    Checks if the `addresses` parameter is a dictionary and logs an error if it's not.
    Checks if the specified `address` is present in the addresses dictionary and logs an error if it's not.
    Checks if the address value is a dictionary and logs an error if it's not.
    Checks if the required keys ('type', 'name', 'protocol', 'port', 'baudrate', 'timeout', 'endline', 'write_timeout',
    'read_termination', 'write_termination', 'bytesize', 'parity') are present in the address dictionary and logs
    an error if any of them are missing.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        addresses (dict): A dictionary containing the configuration addresses.
        address (str): The address to be checked.

    Returns:
        None

    Raises:
        quit(): Exits the program if any errors are encountered.
    """
        
    if not isinstance(addresses, dict):
        logger.critical('Instrument address in config file is not a dictionary! Please edit config file with correct address')
        quit()
    if address not in addresses:
        logger.critical('Instrument address in config file not found! Please edit config file with correct address')
        quit()
    if not isinstance(addresses[address], dict):
        logger.critical('Address {} is not a dictionary! Please edit config file with correct address'.format(address))
        quit()
    if 'type' not in addresses[address]:
        logger.critical('Address {} does not have a type! Please edit config file with correct address'.format(address))
        quit()
    if 'name' not in addresses[address]:
        logger.critical('Address {} does not have a name! Please edit config file with correct address'.format(address))
        quit()
    if 'protocol' not in addresses[address]:
        logger.critical('Address {} does not have a protocol! Please edit config file with correct address'.format(address))
        quit()
    if 'port' not in addresses[address]:
        logger.critical('Address {} does not have a port! Please edit config file with correct address'.format(address))
        quit()
    if 'baudrate' not in addresses[address]:
        logger.critical('Address {} does not have a baudrate! Please edit config file with correct address'.format(address))
        quit()
    if 'timeout' not in addresses[address]:
        logger.critical('Address {} does not have a timeout! Please edit config file with correct address'.format(address))
        quit()
    if 'endline' not in addresses[address]:
        logger.critical('Address {} does not have an endline! Please edit config file with correct address'.format(address))
        quit()
    if 'write_timeout' not in addresses[address]:
        logger.critical('Address {} does not have a write_timeout! Please edit config file with correct address'.format(address))
        quit()
    if 'read_termination' not in addresses[address]:
        logger.critical('Address {} does not have a read_termination! Please edit config file with correct address'.format(address))
        quit()
    if 'write_termination' not in addresses[address]:
        logger.critical('Address {} does not have a write_termination! Please edit config file with correct address'.format(address))
        quit()
    if 'bytesize' not in addresses[address]:
        logger.critical('Address {} does not have a bytesize! Please edit config file with correct address'.format(address))
        quit()
    if 'parity' not in addresses[address]:
        logger.critical('Address {} does not have a parity! Please edit config file with correct address'.format(address))
        quit()

def check_missing_params(logger, my_dict, my_list):
    """
    Check for missing measurement parameters in a dictionary.

    Compares the items in `my_list` against the keys in `my_dict` to find any missing items.
    If no missing items are found, logs an informational message.
    If only one missing item is found, logs a critical message with the missing item and exits the program.
    If multiple missing items are found, logs a critical message with a list of missing items and exits the program.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        my_dict (dict): The dictionary containing the measurement parameters.
        my_list (list): The list of measurement parameters to check.

    Returns:
        None

    Raises:
        quit(): Exits the program if missing items are found.
    """

    missing_items = [item for item in my_list if item not in my_dict]

    if len(missing_items) == 0:
        logger.info('All measurement parameters found!')
    elif len(missing_items) == 1:
        logger.critical(f'Missing measurement parameter: {missing_items[0]}')
        quit()
    else:
        logger.critical('Missing measurement parameters:')
        for item in missing_items:
            logger.critical(f'- {item}')
        quit()

def create_list(logger, condition_values):
    """
    Create a list of values based on the provided condition values.

    Takes a list `condition_values` containing the cycle type, maximum value, minimum value, and step size.
    Generates one-way lists (`listp_oneway` and `listn_oneway`) using `np.arange`.
    Combines the one-way lists with maximum and minimum values to create symmetric lists (`listp` and `listn`)
    using `np.concatenate`.
    Constructs a dictionary `cycle_dict` with cycle types as keys and corresponding lists as values.
    Returns the list of values based on the specified cycle type from the `cycle_dict`.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        condition_values (list): A list containing the cycle type, maximum value, minimum value, and step size.

    Returns:
        numpy.ndarray: The generated list of values based on the specified cycle type.

    Raises:
        KeyError: If an invalid cycle type is provided.
        quit(): Exits the program if an invalid cycle type is provided or if an unknown error occurs.
    """

    cycle, max_value, min_value, step = condition_values
    listp_oneway = np.arange(0, max_value, step)
    listn_oneway = np.arange(0, min_value, -step)
    listp = np.concatenate((listp_oneway, np.array([max_value]), np.flip(listp_oneway)))
    listn = np.concatenate((listn_oneway, np.array([min_value]), np.flip(listn_oneway)))

    try:
        cycle_dict = {'+': listp, '-': listn, '+-': np.concatenate((listp, listn)), '-+': np.concatenate((listn, listp))}
        return cycle_dict[cycle]
    except KeyError:
        logger.critical('Invalid cycle type!')
        quit()
    except Exception as e:
        logger.critical(f'Unknown error occurred: {e}')
        quit()

def get_datetime():
    """
    Get the current date and time as a formatted string.

    Retrieves the current timestamp using `time.time()`.
    Converts the timestamp to a `datetime` object using `datetime.datetime.fromtimestamp`.
    Formats the `datetime` object as a string in the format '%Y-%m-%d %H:%M:%S:%f'.
    Returns the formatted string representing the current date and time.

    Returns:
        str or None: The formatted string representing the current date and time.
        Returns None if an error occurs.
    """

    try:
        return datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S:%f')
    except Exception as e:
        print('Error in get_datetime(): {}'.format(e))
        return None

def get_instrument(logger, inst_type, address, **kwargs):
    """
    Get an instrument object based on the provided instrument type and address.

    Initializes an instrument of the specified `inst_type` at the given `address`.
    Logs an informational message indicating the instrument type and address.
    Attempts to import the corresponding module and create an instance of the instrument class using the `import_module` function.
    If successful, returns the instrument instance initialized with the provided logger, address, and additional arguments.
    If an exception occurs during the initialization, logs an error message and returns None.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        inst_type (str): The type of instrument to be initialized.
        address (str): The address of the instrument.
        **kwargs: Additional keyword arguments to be passed to the instrument initialization.

    Returns:
        object or None: An instance of the instrument class if initialization is successful.
        None if an error occurs during initialization.

    """

    logger.info(f'Initializing {inst_type} at {address}')
    try:
        instrument = import_module(logger, 'instrument', inst_type)
        return instrument(logger, address, **kwargs)
    except Exception as e:
        logger.error(f'Could not initialize {inst_type} at {address}: {e}')
        return None

def import_module(logger, type, inst_type=None, measurement_type=None):
    """
    Import a module and retrieve a class object.

    Imports the specified module and retrieves the corresponding class object.
    The `type` parameter determines whether it's an 'instrument' or 'measurement' module.
    Constructs the file name and object name based on the `type`, `inst_type`, and `measurement_type`.
    Returns the imported class object.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        type (str): The type of module to import. Can be 'instrument' or 'measurement'.
        inst_type (str): The type of instrument. Required if `type` is 'instrument'.
        measurement_type (str): The type of measurement. Required if `type` is 'measurement'.

    Returns:
        object: The imported class object.

    Raises:
        ModuleNotFoundError: If the specified module is not found.
        AttributeError: If the specified class is not found within the module.
        quit(): Exits the program if an error occurs during module or class import.
    """

    file_name, obj_name = '', ''
    
    if type == 'instrument':
        file_name = f'instruments.{inst_type}'
        obj_name = 'Instrument'
    elif type == 'measurement':
        file_name = f'measurements.{measurement_type}'
        obj_name = 'Measurement'
    else:
        logger.critical('Invalid type passed to import_module!')
        quit()

    try:
        file = importlib.import_module(file_name)
        logger.debug(f'{file_name} module imported')
        obj = getattr(file, obj_name)
        logger.debug(f'{obj_name} class imported')
        return obj
    except ModuleNotFoundError as e:
        logger.critical(f'Error importing module {file_name}: {e}')
        quit()
    except AttributeError as e:
        logger.critical(f'Error importing class {obj_name} from module {file_name}: {e}')
        quit()
    except Exception as e:
        logger.critical(f'Unexpected error while importing {file_name} or {obj_name}: {e}')
        quit()

def menu(logger, console, type):
    """
    Display a menu and prompt for user selection.

    Displays a menu based on the specified `type` parameter.
    Constructs the menu name, message, and choices based on the menu type.
    Prints the menu name using `print_newlines` and `console.print`.
    Prompts the user for menu selection using `inquirer.prompt`.
    Logs the selected choice and returns it.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        console (Console): A console object for rich text output.
        type (str): The type of menu to display. Can be 'main' or 'measurements'.

    Returns:
        str: The user-selected choice from the menu.

    Raises:
        quit(): Exits the program if an invalid menu type is provided.
    """

    name, message, choices = '', '', []
    
    if type == 'main':
        name = 'Main Menu'
        message = 'Select an option:'
        choices = ['Exit', 'Select Measurement', 'Print Available Addresses', 'Edit Configuration']
    elif type == 'measurements':
        try:
            list_dir = os.listdir('measurements')
            measurement_types = [filename.split('.')[0] for filename in list_dir if filename.endswith('.py')]
            logger.debug(f'{len(measurement_types)} measurement types found')
        except FileNotFoundError:
            logger.critical('No measurement types found!')
            quit()
        
        name = 'Measurement Selection Menu'
        message = 'Select a measurement:'
        choices = measurement_types
        choices.append('Back')
    else:
        logger.critical('Invalid menu type passed to menu function!')
        quit()

    print_newlines(1)
    console.print(Panel.fit(f"[bold]{name}[/bold]", border_style="green"))
    menu_options = [inquirer.List('choice', message=message, choices=choices)]
    answer = inquirer.prompt(menu_options)
    logger.debug(f'Returning answer: {answer}')
    return answer['choice']

def ask_for_comment(logger):
    """
    Prompt the user to enter a comment.

    Displays a prompt asking the user for a comment.
    Validates that the comment is not empty.
    Logs the entered comment and returns it.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.

    Returns:
        str: The comment entered by the user.

    """
    
    print_newlines(2)
    comment = [inquirer.Text('comment', message="What would you like to comment?", validate=lambda _, x: len(x) > 0),]
    answer = inquirer.prompt(comment)
    while answer is None:
        answer = inquirer.prompt(comment)
    logger.debug(f'Comment: {answer["comment"]}')
    return answer["comment"]

def repeat_measurement(logger, console):
    """
    Prompt the user to repeat a measurement.

    Displays a prompt asking the user if they would like to repeat the measurement.
    Logs the user's answer and returns it.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        console (Console): A console object for rich text output.

    Returns:
        bool: True if the user chooses to repeat the measurement, False otherwise.

    """
    
    print_newlines(2)
    comment = [inquirer.Confirm('repeat', message="Would you like to repeat the measurement?")]
    answer = inquirer.prompt(comment)
    logger.debug(f'Repeat: {answer}')
    return answer['repeat']

def exit(logger, inst):
    """
    Exit the program and close the instrument.

    Logs an informational message indicating that the program is exiting.
    Attempts to close the instrument using the `close` method of the provided `inst` object.
    Logs a message indicating whether the instrument was successfully closed or not.
    Exits the program.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        inst (object): The instrument object to be closed.

    Returns:
        None

    """
    
    logger.info('Exiting program...') 
    try:
        inst.close(logger)
        logger.info('Instrument closed')
    except:
        logger.info('Failed to close instrument')


# Important Classes ---------------------------------------------------------------


class FileSave:
    """
    Class for saving data and configuration to an Excel file.

    The `FileSave` class provides methods for creating a new Excel file,
    saving configuration data to the 'config' sheet, saving headers to the 'data' sheet,
    and saving results to the 'data' sheet.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        path (str): The path where the file will be saved.
        sample (str): The sample name.
        device (str): The device name.

    Attributes:
        path (str): The full path of the file.
        file_name (str): The name of the file.
        file_path (str): The full path of the file including the file name.
        wb (openpyxl.Workbook): The workbook object for the Excel file.

    Methods:
        save_config(logger, config, measurement_type): Save the configuration data to the 'config' sheet.
        save_headers(logger, headers_list): Save the headers to the 'data' sheet.
        save_result(logger, result): Save a result row to the 'data' sheet.
    """

    def __init__(self, logger, path, sample, device):

        """
        Initialize a new Excel file.

        Creates a new Excel file with separate 'config' and 'data' sheets.
        Sets up the file path and name based on the provided `path`, `sample`, and `device`.
        Saves the empty workbook to the specified file path.

        Args:
            logger (logging.Logger): The logger object to use for logging messages.
            path (str): The path where the file will be saved.
            sample (str): The sample name.
            device (str): The device name.

        Returns:
            None
        """
        self.path = os.path.join(path, sample, device)
        mkdir(logger, self.path)

        self.file_name = get_filename(self.path, sample, device)
        self.file_path = os.path.join(self.path, self.file_name)
        self.wb = openpyxl.Workbook()
        self.wb.create_sheet('config')
        self.wb.create_sheet('data')
        self.wb.remove(self.wb['Sheet'])
        self.wb.save(self.file_path)

    def save_config(self, logger, config, measurement_type):
        """
        Save the configuration data to the 'config' sheet.

        Writes the configuration data to the 'config' sheet of the Excel file.
        The `config` parameter is expected to be a dictionary with sections as keys and section data as values.
        Each section is written as a separate block in the sheet.

        Args:
            logger (logging.Logger): The logger object to use for logging messages.
            config (dict): The configuration data.
            measurement_type (str): The type of measurement.

        Returns:
            None
        """
        self.ws = self.wb['config']

        sections = ['sample', 'instrument', measurement_type]
        for section in sections:
            self.ws.append([section])
            for key, value in config.get(section).items():
                self.ws.append([key, value])
            self.ws.append([])
        self.wb.save(self.file_path)
        logger.debug('Configuration file saved in config sheet')

    def save_headers(self, logger, headers_list):
        """
        Save the headers to the 'data' sheet.

        Writes the headers to the first row of the 'data' sheet.
        The `headers_list` parameter is expected to be a list of header names.

        Args:
            logger (logging.Logger): The logger object to use for logging messages.
            headers_list (list): The list of header names.

        Returns:
            None
        """
        self.ws = self.wb['data']
        for i in range(len(headers_list)):
            self.ws.cell(row=1, column=i + 1).value = headers_list[i]
        self.wb.save(self.file_path)
        logger.debug('Headers saved in data sheet')

    def save_result(self, logger, result):
        
        """
        Save a result row to the 'data' sheet.

        Appends a result row to the 'data' sheet.
        The `result` parameter is expected to be a list of values representing a row of results.

        Args:
            logger (logging.Logger): The logger object to use for logging messages.
            result (list): The list of values representing a result row.

        Returns:
            None
        """
            
        self.ws = self.wb['data']
        row = self.ws.max_row + 1
        for i in range(len(result)):
            try:
                self.ws.cell(row=row, column=i + 1).value = result[i]
            except IndexError:
                logger.error('IndexError: result row is too long')
                raise
        self.wb.save(self.file_path)
        logger.debug('result saved in data sheet')

class Logbook:
    """
    Class for managing a logbook in an Excel file.

    ...

    Methods:
        __init__(self, logger, path, sample): Initialize a new logbook.
        save_log(self, logger, date, file, comment): Save a log entry to the logbook.
    """

    def __init__(self, logger, path, sample):
        """
        Initialize a new logbook.

        Creates a new logbook Excel file if it doesn't exist.
        If the logbook file already exists, it loads the existing file.
        Sets up the file path and name based on the provided `path` and `sample`.
        Creates a 'logbook' sheet in the workbook and sets the headers.
        Saves the workbook to the specified file path.

        Args:
            logger (logging.Logger): The logger object to use for logging messages.
            path (str): The path where the logbook file will be saved.
            sample (str): The sample name.

        Returns:
            None
        """

        self.path = os.path.join(path, sample)
        mkdir(logger, self.path)

        self.file_path = os.path.join(self.path, 'logbook.xlsx')

        if path_exists(self.file_path):
            self.wb = openpyxl.load_workbook(self.file_path)
            logger.debug('Logbook file already exists')
        else:
            self.wb = openpyxl.Workbook()
            self.wb.create_sheet('logbook')
            self.wb.remove(self.wb['Sheet'])

            self.ws = self.wb['logbook']
            self.ws.title = 'logbook'
            headers_list = ['Date', 'File', 'Comment']
            for i in range(len(headers_list)):
                self.ws.cell(row=1, column=i + 1).value = headers_list[i]

            self.wb.save(self.file_path)
            logger.debug('Logbook file created')

    def save_log(self, logger, date, file, comment):
        """
        Save a log entry to the logbook.

        Appends a log entry to the 'logbook' sheet of the logbook Excel file.
        Writes the provided `date`, `file`, and `comment` to the respective columns.

        Args:
            logger (logging.Logger): The logger object to use for logging messages.
            date (str): The date of the log entry.
            file (str): The file associated with the log entry.
            comment (str): The comment for the log entry.

        Returns:
            None
        """

        if not len(date) > 0:
            logger.warning('No date provided for log entry.')
            return
        if not len(file) > 0:
            logger.warning('No file provided for log entry.')
            return
        if not len(comment) > 0:
            logger.warning('No comment provided for log entry.')
            return
        try:
            self.ws = self.wb['logbook']
            row = self.ws.max_row + 1
            self.ws.cell(row=row, column=1).value = date
            self.ws.cell(row=row, column=2).value = file
            self.ws.cell(row=row, column=3).value = comment
            self.wb.save(self.file_path)
            logger.debug('Log saved in logbook sheet')
        except Exception as e:
            logger.error(f'Error saving log entry: {e}')

class MeasurementThread(QThread):
    """
    QThread subclass for performing measurements in a separate thread.

    ...

    Signals:
        update_data: Signal emitted to update data in the GUI.
        clear_plots: Signal emitted to clear plots in the GUI.
        close_window: Signal emitted to request the closing of the GUI window.

    Methods:
        __init__(self, logger, inst, meas, config, filesave): Initialize the MeasurementThread object.
        run(self): Run the measurement process in a separate thread.

    Raises:
        ValueError: If n_cycles is not set or is not a positive integer.
    """

    update_data = pyqtSignal(object, list)
    clear_plots = pyqtSignal(list, list)
    close_window = pyqtSignal()

    def __init__(self, logger, inst, meas, config, filesave):
        """
        Initialize the MeasurementThread object.

        Args:
            logger (logging.Logger): The logger object to use for logging messages.
            inst (object): The instrument object for performing measurements.
            meas (object): The measurement object for performing measurements.
            config (dict): The configuration data for the measurement.
            filesave (object): The FileSave object for saving measurement results.

        Raises:
            ValueError: If n_cycles is not set or is not a positive integer.

        Returns:
            None
        """

        super().__init__()
        self.logger = logger
        self.inst = inst
        self.meas = meas
        self.config = config
        self.filesave = filesave
        self.filesave.save_config(logger, config, meas.name)
        self.filesave.save_headers(logger, self.meas.headers)
        self.n_cycles = config.get(meas.name).get('n_cycles')
        self.n_vals = len(self.meas.vals)

    def run(self):
        """
        Run the measurement process in a separate thread.

        Executes the measurement process with the provided instrument, measurement, configuration,
        and FileSave objects.
        Emits signals to update data in the GUI, clear plots, and request the closing of the GUI window.

        Returns:
            None
        """
        self.inst.set_output_state(self.logger,'ON')
        self.inst.write(self.logger, 'INIT:IMM')
        

        with Progress() as progress:
            value_task = progress.add_task(f"[purple] Current value: None", total=self.n_vals) 
            cycle_task = progress.add_task(f"[blue]Cycle 0/{self.n_cycles}", total=self.n_cycles)
            for cycle in range(1, self.n_cycles + 1):
                for val in self.meas.vals:
                    results = self.meas.measure_val(self.logger, self.inst, val)
                    self.filesave.save_result(self.logger, results[0])
                    self.update_data.emit(self.meas, results)
                    progress.update(value_task, advance=1, description=f"[purple]Current value: {val}")
                self.clear_plots.emit(self.meas.plot_grid,self.meas.plot_clear)
                progress.update(value_task, advance=-self.n_vals, description=f"[purple]Value 0/{self.n_vals}")
                progress.update(cycle_task, advance=1, description=f"[blue]Cycle {cycle}/{self.n_cycles}")
            self.inst.set_output_state(self.logger,'OFF')
        self.logger.info('Measurement finished! Please close the window to continue.')

class MeasurementWindow(QMainWindow):
    """
    QMainWindow subclass for displaying measurement results and plots.

    ...

    Signals:
        update_data: Signal emitted to update data in the GUI.
        clear_plots: Signal emitted to clear plots in the GUI.
        close_window: Signal emitted to request the closing of the GUI window.

    Methods:
        __init__(self, logger, inst, meas, config, filesave): Initialize the MeasurementWindow object.
        initUI(self, meas): Initialize the user interface of the window.
        update_data(self, meas, results): Update the displayed data and plots with the measurement results.
        clear_plots(self, plot_grid, plot_clear): Clear the specified plots in the GUI.
        close_window(self): Close the measurement window.

    """

    def __init__(self, logger, inst, meas, config, filesave):
        """
        Initialize the MeasurementWindow object.

        Args:
            logger (logging.Logger): The logger object to use for logging messages.
            inst (object): The instrument object for performing measurements.
            meas (object): The measurement object for performing measurements.
            config (dict): The configuration data for the measurement.
            filesave (object): The FileSave object for saving measurement results.

        Returns:
            None
        """

        super().__init__()
        self.initUI(meas)
        self.measure = MeasurementThread(logger, inst, meas, config, filesave)
        self.measure.update_data.connect(self.update_data)
        self.measure.clear_plots.connect(self.clear_plots)
        self.measure.close_window.connect(self.close)
        self.measure.start()

    def initUI(self, meas):
        """
        Initialize the user interface of the window.

        Args:
            meas (object): The measurement object.

        Returns:
            None
        """

        self.setWindowTitle(meas.name)
        self.setGeometry(100, 100, 800, 600)

        self.tab_widget = QTabWidget(self)
        plot_tab = QWidget()
        plot_layout = QVBoxLayout(plot_tab)
        self.plot_grid = QGridLayout()
        [rows, cols] = meas.plot_grid
        self.data_plots = [[[] for _ in range(cols)] for _ in range(rows)]
        for row in range(rows):
            for col in range(cols):
                plot_widget = pg.PlotWidget()
                plot_widget.setTitle(meas.plot_titles[row][col])
                plot_widget.showGrid(x=True, y=True)
                plot_widget.setLabel('bottom', meas.plot_labels[row][col][0])
                plot_widget.setLabel('left', meas.plot_labels[row][col][1])
                self.plot_grid.addWidget(plot_widget, row, col)

        plot_layout.addLayout(self.plot_grid)
        plot_tab.setLayout(plot_layout)
        self.tab_widget.addTab(plot_tab, "Plots")

        data_tab = QWidget()
        data_layout = QVBoxLayout(data_tab)
        self.data_table = QTableWidget()
        self.data_table.setRowCount(0)
        self.data_table.setColumnCount(len(meas.headers))
        self.data_table.setHorizontalHeaderLabels(meas.headers)
        data_layout.addWidget(self.data_table)
        data_tab.setLayout(data_layout)
        self.tab_widget.addTab(data_tab, "Data")

        self.setCentralWidget(self.tab_widget)

    def update_data(self, meas, results):
        """
        Update the displayed data and plots with the measurement results.

        Args:
            meas (object): The measurement object.
            results (list): The measurement results.

        Returns:
            None
        """

        [result, result_plots] = results
        self.data_table.insertRow(0)
        for col in range(len(meas.headers)):
            item = QTableWidgetItem(str(result[col]))
            self.data_table.setItem(0, col, item)

        for row in range(meas.plot_grid[0]):
            for col in range(meas.plot_grid[1]):
                self.data_plots[row][col].append(result_plots[row][col])
                plot_widget = self.plot_grid.itemAtPosition(row, col).widget()
                plot_widget.plot(
                    np.array(self.data_plots[row][col])[:,0],
                    np.array(self.data_plots[row][col])[:,1], 
                    clear=True)
        
    def clear_plots(self, plot_grid, plot_clear):
        """
        Clear the specified plots in the GUI.

        Args:
            plot_grid (list): The grid layout of the plots.
            plot_clear (list): The boolean flags indicating which plots to clear.

        Returns:
            None
        """    

        for row in range(plot_grid[0]):
            for col in range(plot_grid[1]):
                if plot_clear[row][col]:
                    self.data_plots[row][col] = []

    def close_window(self):
        """
        Close the measurement window.

        Returns:
            None
        """
                
        self.close()

# Main Functions ------------------------------------------------------------------


def start_gui(logger, inst, meas, config, filesave):
    """
    Start the graphical user interface for the measurement application.

    Args:
        logger (logging.Logger): The logger object to use for logging messages.
        inst (object): The instrument object for performing measurements.
        meas (object): The measurement object for performing measurements.
        config (dict): The configuration data for the measurement.
        filesave (object): The FileSave object for saving measurement results.

    Returns:
        None
    """

    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create("Fusion"))
    app.setStyleSheet("QMainWindow::title {background-color: #333333;}")
    window = MeasurementWindow(logger, inst, meas, config, filesave)
    window.show()
    try:
        app.exec()
    except Exception as e:
        logger.error('Error in GUI: %s' % str(e))

def main():
    """
    Main function for running the measurement application.

    Initializes the necessary components, displays the main menu, handles user input,
    and performs the selected actions based on the user's choices.

    Returns:
        None
    """
    
    install()
    console = Console()
    logger = verbose_debug(False)
    running = True

    clear_terminal()
    splash_screen(console)

    config = read_config(logger)

    addresses = get_available_addresses()
    print_available_addresses(logger, console, addresses)

    inst_class = import_module(logger=logger, type='instrument', inst_type=config.get('instrument').get('model'))
    inst = inst_class(logger, config)
    inst.reset(logger)
    logger.info('Instrument loaded')
    atexit.register(exit, logger, inst)

    while running:
        option = menu(logger, console, 'main')
        

        match option:
            case 'Exit':
                running = False

            case 'Print Available Addresses':
                print_available_addresses(logger, console, get_available_addresses())

            case 'Edit Configuration':
                open_config(logger)

            case 'Select Measurement':
                config = read_config(logger)
                option2 = menu(logger, console, 'measurements')
                if option2 == 'Back': continue

                meas_class = import_module(logger=logger, type='measurement', measurement_type=option2)
                meas = meas_class(logger, config, inst)

                filesave = FileSave(logger, config.get('sample').get('path'), config.get('sample').get('name'), config.get('sample').get('device'))

                check_missing_params(logger, meas.params[meas.name], meas.nparams)
                logger.info('All parameters present')
                start_gui(logger, inst, meas, config, filesave)

                path = config.get('sample').get('path')
                sample_name = config.get('sample').get('name')
                sample_device = config.get('sample').get('device') 
                
                logbook = Logbook(logger,path, sample_name)
                
                logger.info('Please enter a comment for the logbook')
                comment = ask_for_comment(logger)
                logbook.save_log(logger, get_datetime(), filesave.file_name, comment)
                logger.info('Logbook saved!')


                option3 = repeat_measurement(logger, console)
                if option3: continue
                else: 
                    running = False

if __name__ == "__main__":
    main()



