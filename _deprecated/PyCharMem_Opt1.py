# Imports
import os
import sys
import time
import logging
import platform
import importlib
import datetime
import openpyxl
import inquirer
import yaml
import pyvisa
import numpy as np
import pyqtgraph as pg
from openpyxl.drawing.image import Image
from art import text2art
from rich.console import Console
from rich.traceback import install
from rich.logging import RichHandler
from rich.progress import Progress
from rich.panel import Panel
from PyQt6.QtCore import QThread, pyqtSignal
from PyQt6.QtWidgets import (QApplication, QMainWindow, QStyleFactory, QTableWidget, 
                              QTableWidgetItem, QVBoxLayout, QWidget, QTabWidget, QGridLayout)

# Global Variables
ProgramInfo = {
    'name': 'PyCharMem',
    'version': '0.0.2',
    'author': 'Ricardo E. Silva',
    'email': 'ricardoedgarsilva@tecnico.ulisboa.pt',
    'description': 'PyCharMem is a Python program that allows you to measure the charge memory of a device.',
    'license': 'GNU General Public License v3.0',
    'url': 'https://github.com/ricardoedgarsilva/PyCharMem'
}

# Common functions
def verbose_debug(verbose: bool) -> logging.Logger:
    if verbose:
        logging.basicConfig(level="NOTSET", format="%(message)s", datefmt="[%X]", handlers=[RichHandler()])
    else:
        logging.basicConfig(level='INFO', format="%(message)s", datefmt="[%X]", handlers=[RichHandler()])

    return logging.getLogger("rich")

def clear_terminal() -> None:
    os.system('cls||clear')

def printnewlines(lines: int) -> None:  
    print(lines * "\n")

def splash_screen(console: Console) -> None:
    print(text2art(ProgramInfo['name']))
    console.print(f"Version: {ProgramInfo['version']}", style='bold blue')
    console.print(f"Author: {ProgramInfo['author']}", style='bold blue')
    printnewlines(5)

def config_open(logger: logging.Logger) -> None:
    try:
        os.system({'Windows': 'start config.yml', 'Darwin': 'open config.yml', 'Linux': 'open config.yml'}.get(platform.system(), ''))
    except:
        logger.critical('Config file not found')
        quit()

def config_read(logger: logging.Logger) -> dict:
    try:
        with open('config.yml', 'r') as file:
            config = yaml.safe_load(file)
        return config
    except:
        logger.critical('Config file not found')
        quit()

def path_exists(path: str) -> bool:
    return os.path.exists(path)

def get_path() -> str:
    return os.path.dirname(os.path.realpath(__file__))

def mkdir(logger: logging.Logger, path: str) -> None:
    if not path_exists(path):
        os.makedirs(path)
        logger.debug(f'Folder {path} created')

def get_index(path: str) -> int:
    return len(os.listdir(path))

def get_filename(path: str, sample: str, device: str) -> str:
    index = get_index(path)
    date = datetime.datetime.now().strftime('%Y%m%d')
    return f'{date}_{sample}_{device}_{index}.xlsx'

def get_available_addresses() -> list:
    rm = pyvisa.ResourceManager()
    addresses = rm.list_resources()
    return addresses

def print_available_addresses(logger: logging.Logger, console: Console, addresses: list) -> None:
    printnewlines(1)
    console.print(Panel.fit("[bold]Available Addresses[/bold]", border_style="green"))
    
    if len(addresses) == 0:
        logger.critical('No addresses found!')
    else:
        for address in addresses:
            console.print(address)

def check_address(logger: logging.Logger, addresses: list, address: str) -> None:
    if address not in addresses:
        logger.critical('Instrument address in config file not found! Please edit config file with correct address')
        quit()

def check_missing_params(logger: logging.Logger, my_dict: dict, my_list: list) -> None:
    missing_items = [item for item in my_list if item not in my_dict]

    if len(missing_items) == 0:
        logger.info('All measurement parameters found!')
    else:
        logger.critical('Missing measurement parameters:')
        for item in missing_items:
            logger.critical(f'- {item}')
        quit()

def create_list(logger: logging.Logger, condition_values: list) -> np.ndarray:
    cycle, max_value, min_value, step = condition_values
    listp_oneway = np.arange(0, max_value, step)
    listn_oneway = np.arange(0, min_value, -step)
    listp = np.concatenate((listp_oneway, np.array([max_value]), np.flip(listp_oneway)))
    listn = np.concatenate((listn_oneway, np.array([min_value]), np.flip(listn_oneway)))

    try:
        cycle_dict = {'+': listp, '-': listn, '+-': np.concatenate((listp, listn)), '-+': np.concatenate((listn, listp))}
        return cycle_dict[cycle]
    except KeyError:
        logger.critical('Invalid cycle type!')
        quit()

def get_datetime() -> str:
    return datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S:%f')

def import_module(logger: logging.Logger, type: str, inst_type=None, measurement_type=None) -> type:
    file_name, obj_name = '', ''
    
    if type == 'instrument':
        file_name = f'instruments.{inst_type}'
        obj_name = 'Instrument'
    elif type == 'measurement':
        file_name = f'measurements.{measurement_type}'
        obj_name = 'Measurement'
    else:
        logger.critical('Invalid type passed to import_module!')
        quit()

    file = importlib.import_module(file_name)
    logger.debug(f'{file_name} module imported')
    obj = getattr(file, obj_name)
    logger.debug(f'{obj_name} class imported')
    return obj

def menu(logger: logging.Logger, console: Console, type: str) -> str:
    name, message, choices = '', '', []
    
    if type == 'main':
        name = 'Main Menu'
        message = 'Select an option:'
        choices = ['Exit', 'Select Measurement', 'Print Available Addresses', 'Edit Configuration']
    elif type == 'measurements':
        try:
            list_dir = os.listdir('measurements')
            measurement_types = [filename.split('.')[0] for filename in list_dir if filename.endswith('.py')]
            logger.debug(f'{len(measurement_types)} measurement types found')
        except:
            logger.critical('No measurement types found!')
            quit()
        
        name = 'Measurement Selection Menu'
        message = 'Select a measurement:'
        choices = measurement_types
        choices.append('Back')
    else:
        logger.critical('Invalid menu type passed to menu function!')
        quit()

    printnewlines(1)
    console.print(Panel.fit(f"[bold]{name}[/bold]", border_style="green"))
    menu_options = [inquirer.List('choice', message=message, choices=choices)]
    answer = inquirer.prompt(menu_options)
    logger.debug(f'Returning answer: {answer}')
    return answer['choice']

def ask_for_comment(logger: logging.Logger) -> str:
    comment = [inquirer.Text('comment', message="What would you like to comment?", validate=lambda _, x: len(x) > 0),]
    answer = inquirer.prompt(comment)
    logger.debug(f'Comment: {answer}')
    return answer['comment']

# Important Classes

class FileSave:
    def __init__(self, logger: logging.Logger, sample: str, device: str):
        # Initializes the class

        # Assign path and make folder if it doesn't exist
        self.path = os.path.join(get_path(), "data", sample, device)
        mkdir(logger, self.path)
        # Create file
        self.file_name = get_filename(self.path, sample, device)
        self.file_path = os.path.join(self.path, self.file_name)
        self.wb = openpyxl.Workbook()
        self.wb.create_sheet('config')
        self.wb.create_sheet('data')
        self.wb.create_sheet('plots')
        self.wb.remove(self.wb['Sheet'])
        self.wb.save(self.file_path)

    def save_config(self, logger: logging.Logger, config: dict, measurement_type: str):
        # Saves the config file in the config sheet

        # Open config sheet and set title
        self.ws = self.wb['config']

        # read config file
        sections = ['sample', 'instrument', measurement_type]
        for section in sections:
            self.ws.append([section])
            for key, value in config.get(section).items():
                self.ws.append([key, value])
            self.ws.append([])
        self.wb.save(self.file_path)
        logger.debug('Configuration file saved in config sheet')

    def save_headers(self, logger: logging.Logger, headers_list: list):
        self.ws = self.wb['data']
        for i in range(len(headers_list)):
            self.ws.cell(row=1, column=i + 1).value = headers_list[i]
        self.wb.save(self.file_path)
        logger.debug('Headers saved in data sheet')

    def save_result(self, logger: logging.Logger, result: list):
        self.ws = self.wb['data']
        row = self.ws.max_row + 1
        for i in range(len(result)):
            self.ws.cell(row=row, column=i + 1).value = result[i]
        self.wb.save(self.file_path)
        logger.debug('result saved in data sheet')

    def save_plots(self, logger: logging.Logger):
        self.ws = self.wb['plots']
        img = Image('temp/plot.png')
        self.ws.add_image(img, 'B2')
        self.wb.save(self.file_path)
        os.remove('temp/plot.png')
        logger.debug('Plot saved in plots sheet')


class Logbook:
    def __init__(self, logger: logging.Logger, sample: str):
        # Initializes the class

        # Assign path and make folder if it doesn't exist
        self.path = os.path.join(get_path(), "data", sample)
        mkdir(logger, self.path)
        # Create file
        self.file_path = os.path.join(self.path, 'logbook.xlsx')

        # Check if file exists
        if path_exists(self.file_path):
            self.wb = openpyxl.load_workbook(self.file_path)
            logger.debug('Logbook file already exists')

        else:
            self.wb = openpyxl.Workbook()
            self.wb.create_sheet('logbook')
            self.wb.remove(self.wb['Sheet'])

            # Create headers
            self.ws = self.wb['logbook']
            self.ws.title = 'logbook'
            headers_list = ['Date', 'File', 'Comment']
            for i in range(len(headers_list)):
                self.ws.cell(row=1, column=i + 1).value = headers_list[i]

            self.wb.save(self.file_path)
            logger.debug('Logbook file created')

    def save_log(self, logger: logging.Logger, date: str, file: str, comment: str):
        # Saves the log in the logbook sheet
        self.ws = self.wb['logbook']
        row = self.ws.max_row + 1
        self.ws.cell(row=row, column=1).value = date
        self.ws.cell(row=row, column=2).value = file
        self.ws.cell(row=row, column=3).value = comment
        self.wb.save(self.file_path)
        logger.debug('Log saved in logbook sheet')


class MeasurementThread(QThread):
    update_data = pyqtSignal(object, list)

    def __init__(self, logger, inst, meas, config):
        super().__init__()

        self.logger = logger
        self.inst = inst
        self.meas = meas

        # Create file save class
        self.filesave = FileSave(logger, config.get('sample').get('name'), config.get('instrument').get('model'))
        self.filesave.save_config(logger, config, meas.name)
        self.filesave.save_headers(logger, self.meas.headers)

        self.n_cycles = config.get(meas.name).get('n_cycles')

    def run(self):

        for cycle in range(1, self.n_cycles + 1):
            for val in self.meas.vals:
                results = self.meas.measure_val(self.logger, self.inst, val)
                self.filesave.save_result(self.logger, results[0])
                self.update_data.emit(self.meas, results)


class MeasurementWindow(QMainWindow):
    def __init__(self, logger, inst, meas, config):
        super().__init__()
        self.initUI(meas)

        self.measure = MeasurementThread(logger, inst, meas, config)
        self.measure.update_data.connect(self.update_data)
        self.measure.start()

    def initUI(self, meas):
        self.setWindowTitle(meas.name)
        self.setGeometry(100, 100, 800, 600)

        # Create a tab widget
        self.tab_widget = QTabWidget(self)

        # Create a Plot tab and add it to the tab widget
        plot_tab = QWidget()
        plot_layout = QVBoxLayout(plot_tab)
        self.plot_grid = QGridLayout()
        [rows, cols] = meas.plot_grid
        self.data_plots = [[[] for _ in range(cols)] for _ in range(rows)]
        for row in range(rows):
            for col in range(cols):
                plot_widget = pg.PlotWidget()
                plot_widget.setTitle(meas.plot_titles[row][col])
                plot_widget.showGrid(x=True, y=True)
                plot_widget.setLabel('bottom', meas.plot_labels[row][col][0])
                plot_widget.setLabel('left', meas.plot_labels[row][col][1])
                self.plot_grid.addWidget(plot_widget, row, col)

        plot_layout.addLayout(self.plot_grid)
        plot_tab.setLayout(plot_layout)
        self.tab_widget.addTab(plot_tab, "Plots")

        # Create a tab and add it to the tab widget
        data_tab = QWidget()
        data_layout = QVBoxLayout(data_tab)
        self.data_table = QTableWidget()
        self.data_table.setRowCount(0)
        self.data_table.setColumnCount(len(meas.headers))
        self.data_table.setHorizontalHeaderLabels(meas.headers)
        data_layout.addWidget(self.data_table)
        data_tab.setLayout(data_layout)
        self.tab_widget.addTab(data_tab, "Data")

        # Add the widget to the main window
        self.setCentralWidget(self.tab_widget)

    def update_data(self, meas, results):

        [result, result_plots] = results
        # Update data table
        self.data_table.insertRow(0)
        for col in range(len(meas.headers)):
                item = QTableWidgetItem(str(result[col]))
                self.data_table.setItem(0, col, item)

        for row in range(meas.plot_grid[0]):
            for col in range(meas.plot_grid[1]):
                self.data_plots[row][col].append(result_plots[row][col])
                plot_widget = self.plot_grid.itemAtPosition(row, col).widget()
                plot_widget.plot(self.data_plots[row][col], clear=True)


# Main Functions ------------------------------------------------------------------

def start_gui(logger: logging.Logger, inst: object, meas: object, config: dict):
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create("Fusion"))
    app.setStyleSheet("QMainWindow::title {background-color: #333333;}")
    window = MeasurementWindow(logger, inst, meas, config)
    window.show()
    app.exec()


def main():
    install()
    console = Console()
    logger = verbose_debug(False)
    running = True

    # Splash Screen
    clear_terminal()
    splash_screen(console)

    # Load config file
    config = config_read(logger)
    logger.info('Configuration file loaded')

    # Print Available Addresses
    addresses = get_available_addresses()
    print_available_addresses(logger, console, addresses)

    # Load Instruments
    inst_class = import_module(logger=logger, type='instrument', inst_type=config.get('instrument').get('model'))
    inst = inst_class(logger, config)
    logger.info('Instrument loaded')

    while running:
        option = menu(logger, console, 'main')

        match option:
            case 'Exit':
                running = False
                logger.info('Exiting program')

            case 'Print Available Addresses':
                print_available_addresses(logger, console, get_available_addresses())

            case 'Edit Configuration':
                config_open(logger)

            case 'Select Measurement':
                # Load Measurements menu and go back to main menu if 'Back' is selected
                option2 = menu(logger, console, 'measurements')
                if option2 == 'Back': continue

                # Import measurement class
                meas_class = import_module(logger=logger, type='measurement', measurement_type=option2)
                meas = meas_class(logger, config, inst)

                # Check if all measurements are present
                check_missing_params(logger, meas.params[meas.name], meas.nparams)
                logger.info('All parameters present')
                logbook = Logbook(logger, config.get('sample').get('name'))

                start_gui(logger, inst, meas, config)

                # Comment
                inst.close(logger)
                comment = ask_for_comment(logger)
                logbook.save_log(logger, get_datetime(), filesave.file_name, comment)


if __name__ == "__main__":
    main()


